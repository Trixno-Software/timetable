from django.http import HttpResponse
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.permissions import IsCoordinator
from apps.timetable.models import Timetable

from .generators import (
    CSVExportGenerator,
    ExcelExportGenerator,
    ExportScope,
    PDFExportGenerator,
)


class TimetableExportView(APIView):
    """Export timetable in various formats"""
    permission_classes = [IsAuthenticated, IsCoordinator]

    def get(self, request, timetable_id):
        try:
            timetable = Timetable.objects.get(id=timetable_id)
        except Timetable.DoesNotExist:
            return Response(
                {"error": "Timetable not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get export parameters
        export_format = request.query_params.get("format", "xlsx").lower()
        scope = request.query_params.get("scope", ExportScope.SCHOOL)
        scope_id = request.query_params.get("scope_id")

        # Validate scope
        if scope not in [ExportScope.SCHOOL, ExportScope.GRADE, ExportScope.SECTION, ExportScope.TEACHER]:
            return Response(
                {"error": "Invalid scope. Must be one of: school, grade, section, teacher"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if scope in [ExportScope.GRADE, ExportScope.SECTION, ExportScope.TEACHER] and not scope_id:
            return Response(
                {"error": f"scope_id is required for {scope} scope"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            if export_format == "xlsx":
                return self._export_excel(timetable, scope, scope_id)
            elif export_format == "csv":
                return self._export_csv(timetable, scope, scope_id)
            elif export_format == "pdf":
                return self._export_pdf(timetable, scope, scope_id)
            else:
                return Response(
                    {"error": "Invalid format. Must be one of: xlsx, csv, pdf"},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response(
                {"error": f"Export failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _export_excel(self, timetable, scope, scope_id):
        generator = ExcelExportGenerator(timetable)
        generator.generate_for_scope(scope, scope_id)
        content = generator.to_bytes()

        filename = f"timetable_{timetable.name}_{scope}.xlsx"
        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _export_csv(self, timetable, scope, scope_id):
        generator = CSVExportGenerator(timetable)
        content = generator.generate_for_scope(scope, scope_id)

        filename = f"timetable_{timetable.name}_{scope}.csv"
        response = HttpResponse(content, content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _export_pdf(self, timetable, scope, scope_id):
        generator = PDFExportGenerator(timetable)
        content = generator.generate_pdf(scope, scope_id)

        filename = f"timetable_{timetable.name}_{scope}.pdf"
        response = HttpResponse(content, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class ExportTemplatesView(APIView):
    """Download Excel templates for import"""
    permission_classes = [IsAuthenticated]

    def get(self, request, template_type):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        if template_type == "teachers":
            ws.title = "Teachers"
            headers = ["teacher_code", "teacher_name", "email", "phone", "subjects"]
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).fill = header_fill
                ws.cell(row=1, column=col).font = header_font

            # Sample data
            ws.append(["T001", "John Doe", "john@school.com", "1234567890", "Mathematics, Physics"])
            ws.append(["T002", "Jane Smith", "jane@school.com", "0987654321", "English, History"])

            filename = "teachers_template.xlsx"

        elif template_type == "assignments":
            ws.title = "Assignments"
            headers = ["grade", "section", "subject", "teacher_code", "weekly_periods"]
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).fill = header_fill
                ws.cell(row=1, column=col).font = header_font

            # Sample data
            ws.append(["10", "A", "Mathematics", "T001", "6"])
            ws.append(["10", "A", "Physics", "T001", "4"])
            ws.append(["10", "B", "English", "T002", "5"])

            filename = "assignments_template.xlsx"

        elif template_type == "sections":
            ws.title = "Sections"
            headers = ["grade", "section", "shift", "room", "capacity"]
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).fill = header_fill
                ws.cell(row=1, column=col).font = header_font

            # Sample data
            ws.append(["10", "A", "Morning", "Room 101", "40"])
            ws.append(["10", "B", "Morning", "Room 102", "40"])

            filename = "sections_template.xlsx"

        else:
            return Response(
                {"error": "Invalid template type"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Set column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 5

        # Save to bytes
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
