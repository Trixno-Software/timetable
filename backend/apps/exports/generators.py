"""
Export generators for PDF and Excel formats.
"""

import io
from collections import defaultdict

from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.academics.models import DayOfWeek, PeriodSlot, PeriodTemplate
from apps.timetable.models import Timetable, TimetableEntry


class ExportScope:
    SCHOOL = "school"
    GRADE = "grade"
    SECTION = "section"
    TEACHER = "teacher"


DAY_NAMES = {
    0: "Monday",
    1: "Tuesday",
    2: "Wednesday",
    3: "Thursday",
    4: "Friday",
    5: "Saturday",
    6: "Sunday",
}


class TimetableExportGenerator:
    """Base class for timetable exports"""

    def __init__(self, timetable: Timetable):
        self.timetable = timetable
        self.entries = list(
            TimetableEntry.objects.filter(timetable=timetable)
            .select_related(
                "section", "section__grade", "subject",
                "teacher", "period_slot", "room"
            )
            .order_by("day_of_week", "period_slot__period_number")
        )

    def get_period_slots(self):
        """Get period slots for the timetable"""
        template = PeriodTemplate.objects.filter(
            branch=self.timetable.branch,
            shift=self.timetable.shift,
            is_active=True,
        ).first()

        if template:
            return list(template.slots.filter(is_break=False).order_by("period_number"))
        return []

    def get_entries_by_section(self, section_id):
        """Get entries organized by day and period for a section"""
        grid = defaultdict(dict)
        for entry in self.entries:
            if str(entry.section_id) == str(section_id):
                grid[entry.day_of_week][entry.period_slot.period_number] = entry
        return grid

    def get_entries_by_teacher(self, teacher_id):
        """Get entries organized by day and period for a teacher"""
        grid = defaultdict(dict)
        for entry in self.entries:
            if str(entry.teacher_id) == str(teacher_id):
                grid[entry.day_of_week][entry.period_slot.period_number] = {
                    "entry": entry,
                    "section": f"{entry.section.grade.name}-{entry.section.name}",
                }
        return grid


class ExcelExportGenerator(TimetableExportGenerator):
    """Generate Excel exports"""

    def __init__(self, timetable: Timetable):
        super().__init__(timetable)
        self.workbook = Workbook()
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate_section_sheet(self, section, sheet=None):
        """Generate a sheet for a section's timetable"""
        if sheet is None:
            sheet = self.workbook.create_sheet(
                f"{section.grade.name}-{section.name}"
            )

        period_slots = self.get_period_slots()
        grid = self.get_entries_by_section(str(section.id))

        # Title
        sheet.merge_cells("A1:H1")
        sheet["A1"] = f"Timetable: {section.grade.name} - Section {section.name}"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A1"].alignment = Alignment(horizontal="center")

        # Headers
        sheet["A3"] = "Day/Period"
        sheet["A3"].fill = self.header_fill
        sheet["A3"].font = self.header_font
        sheet["A3"].border = self.border

        col = 2
        for slot in period_slots:
            cell = sheet.cell(row=3, column=col)
            cell.value = f"P{slot.period_number}\n{slot.start_time.strftime('%H:%M')}-{slot.end_time.strftime('%H:%M')}"
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            sheet.column_dimensions[get_column_letter(col)].width = 15
            col += 1

        # Data rows
        row = 4
        for day in range(6):  # Mon-Sat
            day_cell = sheet.cell(row=row, column=1)
            day_cell.value = DAY_NAMES.get(day, "")
            day_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            day_cell.font = Font(bold=True)
            day_cell.border = self.border

            col = 2
            for slot in period_slots:
                cell = sheet.cell(row=row, column=col)
                entry = grid.get(day, {}).get(slot.period_number)
                if entry:
                    cell.value = f"{entry.subject.short_name or entry.subject.name}\n({entry.teacher.first_name[:1]}. {entry.teacher.last_name})"
                else:
                    cell.value = "-"
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                col += 1
            row += 1

        sheet.column_dimensions["A"].width = 12

        return sheet

    def generate_teacher_sheet(self, teacher, sheet=None):
        """Generate a sheet for a teacher's timetable"""
        if sheet is None:
            sheet = self.workbook.create_sheet(
                f"{teacher.first_name} {teacher.last_name}"[:31]
            )

        period_slots = self.get_period_slots()
        grid = self.get_entries_by_teacher(str(teacher.id))

        # Title
        sheet.merge_cells("A1:H1")
        sheet["A1"] = f"Timetable: {teacher.full_name} ({teacher.employee_code})"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A1"].alignment = Alignment(horizontal="center")

        # Headers
        sheet["A3"] = "Day/Period"
        sheet["A3"].fill = self.header_fill
        sheet["A3"].font = self.header_font
        sheet["A3"].border = self.border

        col = 2
        for slot in period_slots:
            cell = sheet.cell(row=3, column=col)
            cell.value = f"P{slot.period_number}\n{slot.start_time.strftime('%H:%M')}-{slot.end_time.strftime('%H:%M')}"
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            sheet.column_dimensions[get_column_letter(col)].width = 15
            col += 1

        # Data rows
        row = 4
        for day in range(6):
            day_cell = sheet.cell(row=row, column=1)
            day_cell.value = DAY_NAMES.get(day, "")
            day_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            day_cell.font = Font(bold=True)
            day_cell.border = self.border

            col = 2
            for slot in period_slots:
                cell = sheet.cell(row=row, column=col)
                data = grid.get(day, {}).get(slot.period_number)
                if data:
                    entry = data["entry"]
                    cell.value = f"{entry.subject.short_name or entry.subject.name}\n{data['section']}"
                else:
                    cell.value = "-"
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                col += 1
            row += 1

        sheet.column_dimensions["A"].width = 12
        return sheet

    def generate_for_scope(self, scope, scope_id=None):
        """Generate Excel based on scope"""
        from apps.academics.models import Grade, Section, Teacher

        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]

        if scope == ExportScope.SECTION:
            section = Section.objects.get(id=scope_id)
            self.generate_section_sheet(section)

        elif scope == ExportScope.GRADE:
            grade = Grade.objects.get(id=scope_id)
            for section in grade.sections.filter(is_active=True):
                self.generate_section_sheet(section)

        elif scope == ExportScope.TEACHER:
            teacher = Teacher.objects.get(id=scope_id)
            self.generate_teacher_sheet(teacher)

        elif scope == ExportScope.SCHOOL:
            # All sections
            sections = Section.objects.filter(
                grade__branch=self.timetable.branch,
                is_active=True
            ).select_related("grade")
            for section in sections:
                self.generate_section_sheet(section)

        # Create summary sheet
        self._create_summary_sheet()

        return self.workbook

    def _create_summary_sheet(self):
        """Create a summary sheet"""
        sheet = self.workbook.create_sheet("Summary", 0)

        sheet["A1"] = "Timetable Export Summary"
        sheet["A1"].font = Font(bold=True, size=16)

        sheet["A3"] = "Timetable Name:"
        sheet["B3"] = self.timetable.name

        sheet["A4"] = "Session:"
        sheet["B4"] = self.timetable.session.name

        sheet["A5"] = "Shift:"
        sheet["B5"] = self.timetable.shift.name

        if self.timetable.season:
            sheet["A6"] = "Season:"
            sheet["B6"] = self.timetable.season.name

        sheet["A7"] = "Status:"
        sheet["B7"] = self.timetable.status

        sheet["A8"] = "Version:"
        sheet["B8"] = self.timetable.current_version

        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 30

    def to_bytes(self):
        """Convert workbook to bytes"""
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()


class CSVExportGenerator(TimetableExportGenerator):
    """Generate CSV exports"""

    def generate_for_scope(self, scope, scope_id=None):
        """Generate CSV content"""
        import csv
        from io import StringIO

        from apps.academics.models import Grade, Section, Teacher

        output = StringIO()
        writer = csv.writer(output)

        period_slots = self.get_period_slots()

        # Header
        header = ["Day"]
        for slot in period_slots:
            header.append(f"Period {slot.period_number}")
        writer.writerow(header)

        if scope == ExportScope.SECTION:
            section = Section.objects.get(id=scope_id)
            grid = self.get_entries_by_section(str(section.id))

            writer.writerow([f"Section: {section.grade.name} - {section.name}"])

            for day in range(6):
                row = [DAY_NAMES.get(day, "")]
                for slot in period_slots:
                    entry = grid.get(day, {}).get(slot.period_number)
                    if entry:
                        row.append(f"{entry.subject.name} ({entry.teacher.full_name})")
                    else:
                        row.append("-")
                writer.writerow(row)

        elif scope == ExportScope.TEACHER:
            teacher = Teacher.objects.get(id=scope_id)
            grid = self.get_entries_by_teacher(str(teacher.id))

            writer.writerow([f"Teacher: {teacher.full_name}"])

            for day in range(6):
                row = [DAY_NAMES.get(day, "")]
                for slot in period_slots:
                    data = grid.get(day, {}).get(slot.period_number)
                    if data:
                        entry = data["entry"]
                        row.append(f"{entry.subject.name} ({data['section']})")
                    else:
                        row.append("-")
                writer.writerow(row)

        output.seek(0)
        return output.getvalue()


class PDFExportGenerator(TimetableExportGenerator):
    """Generate PDF exports using HTML templates"""

    def generate_html(self, scope, scope_id=None):
        """Generate HTML content for PDF"""
        from apps.academics.models import Grade, Section, Teacher

        period_slots = self.get_period_slots()
        context = {
            "timetable": self.timetable,
            "period_slots": period_slots,
            "day_names": DAY_NAMES,
            "days": range(6),
        }

        if scope == ExportScope.SECTION:
            section = Section.objects.select_related("grade").get(id=scope_id)
            grid = self.get_entries_by_section(str(section.id))
            context["section"] = section
            context["grid"] = dict(grid)
            context["title"] = f"{section.grade.name} - Section {section.name}"
            template = "exports/section_timetable.html"

        elif scope == ExportScope.TEACHER:
            teacher = Teacher.objects.get(id=scope_id)
            grid = self.get_entries_by_teacher(str(teacher.id))
            context["teacher"] = teacher
            context["grid"] = dict(grid)
            context["title"] = f"{teacher.full_name}"
            template = "exports/teacher_timetable.html"

        elif scope == ExportScope.GRADE:
            grade = Grade.objects.get(id=scope_id)
            sections_data = []
            for section in grade.sections.filter(is_active=True):
                grid = self.get_entries_by_section(str(section.id))
                sections_data.append({
                    "section": section,
                    "grid": dict(grid),
                })
            context["grade"] = grade
            context["sections_data"] = sections_data
            context["title"] = f"Grade {grade.name}"
            template = "exports/grade_timetable.html"

        else:  # SCHOOL
            grades_data = []
            grades = Grade.objects.filter(
                branch=self.timetable.branch, is_active=True
            ).order_by("order")
            for grade in grades:
                sections_data = []
                for section in grade.sections.filter(is_active=True):
                    grid = self.get_entries_by_section(str(section.id))
                    sections_data.append({
                        "section": section,
                        "grid": dict(grid),
                    })
                grades_data.append({
                    "grade": grade,
                    "sections_data": sections_data,
                })
            context["grades_data"] = grades_data
            context["title"] = f"School Timetable - {self.timetable.branch.name}"
            template = "exports/school_timetable.html"

        return render_to_string(template, context)

    def generate_pdf(self, scope, scope_id=None):
        """Generate PDF bytes"""
        try:
            from weasyprint import HTML
            html_content = self.generate_html(scope, scope_id)
            pdf = HTML(string=html_content).write_pdf()
            return pdf
        except ImportError:
            # WeasyPrint not available, return HTML
            return self.generate_html(scope, scope_id).encode()
